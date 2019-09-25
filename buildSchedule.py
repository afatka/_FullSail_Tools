#! #!/usr/bin/env python3

# Modeling Department Schedule Builder
# Input > Master Schedule
# Output > Schedule containing classes in list

import os, sys, pprint, datetime, warnings, operator
import openpyxl, xlrd, pyperclip #external modules / dependencies required!


def print_usage():
	log('print usage')
	print(''' Build Schedule usage
		Send file with path to an xls or xlsx file to a schedule file. 
		Department schedule will be copied to the clipboard for pasting into a spreadsheet (i.e. google sheets)

		Required Modules: openpyxl, xlrd, pyperclip''')

def buildSchedule(call_argv):
	log('Building Schedule...')
	status('Building Schedule')

	# classes_of_interest = ['3DF', 'MCR', 'CDC', 'PRM'] # 'STE', 'CA4' , '3DA1'
	classes_of_interest = ['MCR']
	element_of_division = ['week1', 'week2', 'week3', 'week4']
	cell_values_to_ignore = ['TBA', 'OFF CAMPUS']
	output_file_name_append = 'Department Schedule'
	cell_of_interest = 5 #index 0 look at row of cells
	cells_to_keep = [0,1,2,3,4,5,8,10] #index 0 cells to capture from schedule row
	department_schedule = {}

	for c in classes_of_interest:
		department_schedule.setdefault(c, [])

	status('Validating Input')
	# log('call_argv: {}'.format(call_argv))
	if len(call_argv) != 2:
		raise InputError('Please include master schedule path as command line argument.')

	master_schedule_path = call_argv[1]

	if not os.path.isfile(master_schedule_path):
		raise InputError('Provided Master Schedule Path is not a file.')

	if os.path.splitext(master_schedule_path)[1] not in ('.xlsx', '.xls'):
		log('split ext: {}'.format(os.path.splitext(master_schedule_path)))
		# if os.path.splitext(master_schedule_path)[1] == '.xls':
		# 	raise InputError('Please convert to the .xlsx file format to process.')
		# else:
		raise InputError('Unsupport file extension:{}'.format(os.path.splitext(master_schedule_path)[1]))

	log('Collect Shedule')
	if master_schedule_path.endswith('.xlsx'):
		log('file is of type: .xlsx')
		collect_from_xlsx(master_schedule_path, department_schedule, cell_of_interest, cells_to_keep, cell_values_to_ignore)
	elif master_schedule_path.endswith('.xls'):
		log('file is of type: .xls')
		collect_from_xls(master_schedule_path, department_schedule, cell_of_interest, cells_to_keep, cell_values_to_ignore)

	pp = pprint.PrettyPrinter(width = 160)
	# pp.pprint(department_schedule)

	log('Format Schedule')
	#separate schedule into weeks
	formatted_department_schedule = format_department_schedule(department_schedule, classes_of_interest, element_of_division)
	# pp.pprint(formatted_department_schedule)

	#build rows in openpyxl workbook

	fluff_weeks(formatted_department_schedule)

	# pp.pprint(formatted_department_schedule)

	build_string_to_copy(formatted_department_schedule, classes_of_interest, element_of_division)

	status('Schedule Building Complete')
	print('')

def build_string_to_copy(schedule_dict, classes_of_interest, element_of_division):
	status('Building string to paste')
	log('Building string to copy')
	long_space = '\t' * 10
	short_space = '\t' * 4
	final_string = '\n' * 2
	for class_title in classes_of_interest:
		for week in element_of_division:
			final_string += class_title + '\t' + week + '\t' + long_space
		final_string += '\n' * 2
		week_iter = 0
		# log('range len: {}'.format(len(schedule_dict[class_title][element_of_division[0]])))
		day_space_check = ''
		add_newline = False
		for i in range(len(schedule_dict[class_title][element_of_division[0]])):
			# log('i: {}'.format(i))
			# log('iter: {}'.format(week_iter))
			for week_title in element_of_division:
				if day_space_check == '':
					day_space_check =  schedule_dict[class_title][week_title][week_iter][0]
				else:
					if day_space_check !=  schedule_dict[class_title][week_title][week_iter][0]:
						day_space_check =  schedule_dict[class_title][week_title][week_iter][0]
						final_string += '\n'
				# log('week_title: {}'.format(week_title))
				# log('week: {}'.format(schedule_dict[class_title][week_title]))
				# log('len: {}'.format(len(schedule_dict[class_title][week_title])))

				if week_iter == len(schedule_dict[class_title][week_title]):
					pp = pprint.PrettyPrinter(width = 160)
					pp.pprint(schedule_dict[class_title][week_title])
				try:
					for elem in schedule_dict[class_title][week_title][week_iter]:
						final_string += str(elem) + '\t'
				except IndexError:
					log('len week in question: \n{}'.format(len(schedule_dict[class_title][week_title])))
					log('week in question:')
					for day in schedule_dict[class_title][week_title]:
						log('{}'.format(day))
					log('week_iter: {}'.format(week_iter))
					log('length of week [0]: {}'.format(len(schedule_dict[class_title][element_of_division[0]])))
					log('week[0]')
					for day in schedule_dict[class_title][element_of_division[0]]:
						log('{}'.format(day))
					raise
				# log('iter: {}'.format(week_iter))
				# log('week: {}'.format(schedule_dict[class_title][week_title]))
				final_string += short_space

			final_string += '\n'
			week_iter += 1
		final_string += '\n'
		
	pyperclip.copy(final_string)
	log('Scheduled Copied.')

def build_rows(class_dict, worksheet_obj):
	log('build rows')

def fluff_weeks(class_dict):
	# log('fluff weeks - redo')
	status('Buffering Weeks')
	pp = pprint.PrettyPrinter(width = 160)
	# log('class_dict:')
	# pp.pprint(class_dict)

	weekdays_list = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

	#TODO - Collect and count days present 
	for class_title in class_dict:
		class_counter_dict = {}
		week_day_list_dict = {}
		max_days = 0
		for week in class_dict[class_title]:
			week_counter_dict = {}
			for day in class_dict[class_title][week]:
				day_key = day[0]
				count = week_counter_dict.get(day_key, 0)
				week_counter_dict[day_key] = count + 1
				class_counter_dict['length'] = len(day)
			# log(week_counter_dict)
			for k in week_counter_dict:
				class_count = class_counter_dict.get(k, 0)
				if week_counter_dict[k] > class_count:
					class_counter_dict[k] = week_counter_dict[k]

	#TODO - Break days into lists of similar day lists
			day_list_dict = {}
			# log('week: {}'.format(class_dict[class_title][week]))
			for day in class_dict[class_title][week]:
				day_key = day[0]
				day_list = day_list_dict.get(day_key, [])
				day_list.append(day)
				day_list_dict[day_key] = day_list
				week_day_list_dict[week] = day_list_dict

			#compensate for empty weeks
			if class_dict[class_title][week] == []:
				# log('class week empty: {}'.format(week))
				week_day_list_dict[week] = day_list_dict

			#find the maximum number of times ANY day is present
			temp_dict = {}
			for key in class_counter_dict:
				if key != 'length':
					temp_dict[key] = class_counter_dict[key]

			try: 
				max_occurance = max(temp_dict.items(), key = operator.itemgetter(1))[1]
			except ValueError as e:
				max_occurance = 0
				log('ValueError: {}'.format(e))
				log('empty week')
			# log('max_occurance: {}'.format(max_occurance))
			if max_occurance > max_days:
				max_days = max_occurance
		# log('max day: {}'.format(max_days))
		# log('Day Counter: {}'.format(class_counter_dict))

		#fluff days to match max iterations of ANY day
		# log('week_day_list_dict: {}'.format(week_day_list_dict))
		for day in weekdays_list:
			# log('Day - {}'.format(day))
			for week in week_day_list_dict:
				# log('Week - {}'.format(week))
				if day in class_counter_dict:
					if day in week_day_list_dict[week]:
						# log('day: {}'.format(day))
						# log('max_days: {}'.format(max_days))
						# log('len dict[week][day]: {}'.format(len(week_day_list_dict[week][day])))
						while len(week_day_list_dict[week][day]) < max_days:
							insert_list = []
							for i in range(class_counter_dict['length']):
								insert_list.append('')
					
							insert_list[0] = day
							insert_list[5] = class_title
							# log('i_list: {}'.format(insert_list))
							week_day_list_dict[week][day].append(insert_list)
					else:
						# log('no days exist!')
						# log('day: {}'.format(day))
						# log('max_days: {}'.format(max_days))
						week_day_list_dict[week][day] = []
						while len(week_day_list_dict[week][day]) < max_days:
							insert_list = []
							for i in range(class_counter_dict['length']):
								insert_list.append('')
							insert_list[0] = day
							insert_list[5] = class_title
							# log('i_list: {}'.format(insert_list))
							week_day_list_dict[week][day].append(insert_list)

				# log('{}: \n{}'.format(week, week_day_list_dict[week]))

		
		# log('max day: {}'.format(max_days))
		# log('Day Counter: {}'.format(class_counter_dict))

		# log('week_day_list_dict: ')
		# pp.pprint(week_day_list_dict)

		#rebuild input dict
		for week in week_day_list_dict:
			replace_list = []
			# log('week: {}'.format(week))
			for day in weekdays_list:
				if day in week_day_list_dict[week]:
					for session in week_day_list_dict[week][day]:
						replace_list.append(session)
			# log('replace_list: \n{}'.format(replace_list))
			class_dict[class_title][week] = replace_list

		# log('class_dict:')
		# pp.pprint(class_dict)

	#TODO - Compare days list to len(most days) -> 
		# - Determine which sections are missing
		# - Add missing sections

	#Todo - rebuild week dictionary

def get_day_index(day_list, weekday, recursion_count = 0):

	# log('day_list: {}'.format(day_list))
	# log('weekday: {}'.format(weekday))

	recursion_count += 1
	if recursion_count > 10:
		# log('Maximum Recursion Reached!')
		# log('day_list: {}'.format(day_list))
		# log('weekday: {}'.format(weekday))
		raise Exception('Recursion Maximum Reached!')

	if len(day_list) < 1:
		# log('day_list: {}'.format(day_list))
		# log('weekday: {}'.format(weekday))
		return 0
		# raise Exception('Day List is empty')

	weekdays_list = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

	try:
		day_index = (len(day_list) - 1) - day_list[::-1].index(weekday)
		# log('index: {}'.format(day_index))
		return day_index
	except ValueError as e:
		# warnings.warn('Day {}.'.format(e))
		prior_day_index = weekdays_list.index(weekday) - 1
		# log('prior_day_index: {}'.format(prior_day_index))
		if prior_day_index < 0:
			# log('IS LESS THAN ZERO')
			return prior_day_index
			# prior_day_index = len(weekdays_list) - 1

		return get_day_index(day_list, weekdays_list[prior_day_index], recursion_count = recursion_count)

#break classes into weeks 
def format_department_schedule(department_schedule, classes_of_interest, element_of_division):
	status('Formatting')
	formatted_department_schedule = {}
	for c in classes_of_interest:
		formatted_department_schedule.setdefault(c, {})
		for week in element_of_division:
			formatted_department_schedule[c].setdefault(week, [])
	for c in classes_of_interest:
		log('c: {}'.format(c))
		week = iter(element_of_division)
		week_num = 0
		previous_day = ''
		for s in department_schedule[c]:
			schedule_date = datetime.date.isocalendar(s[1])[1]
			# log('week: {}'.format(datetime.date.isocalendar(s[1])[1]))
			log('previous_day: {}'.format(previous_day))
			if s[0] == 'Sun' and previous_day != 'Sun':
				week_num = 0
			if schedule_date != week_num:
				log('updating week')
				if previous_day != 'Sun':
					current_week = next(week)
				week_num = schedule_date 
				log('{}'.format(current_week))

			log('current_week: {}'.format(current_week))
			log('s?:{}'.format(s))
			previous_day = s[0]
			formatted_department_schedule[c][current_week].append(s)

	return formatted_department_schedule

#collect shedule info from xls and convert it to openpyxl format place cell objects in dictionary
def collect_from_xls(master_schedule_path, department_schedule, cell_of_interest, cells_to_keep, cell_values_to_ignore):
	status('Parsing xls')
	master_schedule = xlrd.open_workbook(master_schedule_path)
	master_sheet = master_schedule.sheet_by_index(0)

	wb = openpyxl.Workbook()
	ws = wb.active

	for row_index in range(1, master_sheet.nrows):
		cell_of_interest_value = master_sheet.row_values(row_index)[cell_of_interest]
		if cell_of_interest_value in department_schedule:
			# log('cell_of_interest_value found! {}'.format(cell_of_interest_value))
			row = master_sheet.row(row_index)
			# log('row: {}'.format(row))
			for index, cell in enumerate(row):
				# log('{}:{} = {}'.format(index, cell, cell.value))
				# log('Cell Type: {}'.format(cell.ctype))
				if cell.ctype == 3:
					new_value = xlrd.xldate_as_tuple(cell.value, master_schedule.datemode)
					new_value = datetime.datetime(*new_value[0:6])
					# log('Date Found: {}'.format(str(new_value)))
				elif cell.ctype == 2:
					new_value = int(cell.value)
				else: 
					new_value = cell.value
				ws.cell(row = row_index, column = index + 1).value = new_value
				# log('cell.value: {}'.format(cell.value))
				# log('ws.cell.value: {}'.format(ws.cell(row = row_index, column = index + 1).value))

			startCellIndex = 'A' + str(row_index)
			endCellIndex = openpyxl.utils.get_column_letter(master_sheet.ncols) + str(row_index)
			# log('slice index: {}:{}'.format(startCellIndex, endCellIndex))

			openpyxl_row = tuple(ws[startCellIndex:endCellIndex])[0]
			# for i, cell in enumerate(openpyxl_row):
			# 	log('{} {}'.format(i, cell.value))
			
			day = []
			for cell in cells_to_keep:
				day.append(openpyxl_row[cell].value)
			# log('day: {}'.format(day))
			# log('dat')
			if not any(x in day for x in cell_values_to_ignore):
				department_schedule[cell_of_interest_value].append(day)	

#collect schedule info from xlsx file place openpyxl cell objects in dictionary
def collect_from_xlsx(master_schedule_path, department_schedule, cell_of_interest, cells_to_keep, cell_values_to_ignore):
	status('Parsing xlsx')
	master_schedule = openpyxl.load_workbook(master_schedule_path)
	master_sheet = master_schedule.active

	# for r in master_sheet.rows[1:]:
	for r in tuple(master_sheet.rows)[1:]:
		if r[cell_of_interest].value in department_schedule:
			# log('r[cell_of_interest].value found! {}'.format(r[cell_of_interest].value))
			day = []
			for cell in cells_to_keep:
				day.append(r[cell].value)
			if not any(x in day for x in cell_values_to_ignore):
				department_schedule[r[cell_of_interest].value].append(day)

def log(message, development = False):
	if development:
		print(message)

def status(message):
	print('\r{}'.format(message), end = '')
	# sys.stdout.flush()
	

class InputError(Exception):
	pass

if len(sys.argv) == 1:
	print_usage()
else:
	buildSchedule(sys.argv)
